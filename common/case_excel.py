#coding:utf-8
import os
from openpyxl import load_workbook

class ExcelCase(object):
	casedic={'caseid':u'用例编号','content':u'测试内容','target':u'测试目的',
			'condition':u'前提条件','step':u'测试步骤','expect':u'预期结果',
			'result':u'测试结果','explain':u'未通过说明','bugid':u'BUG ID','runflag':u'Run_Flag'}

	def __init__(self,excel_file):
		self.excel_file=excel_file

	def excel_read(self):
		excel_rows=[]
		wb=load_workbook(self.excel_file)
		for sheet_name in wb.get_sheet_names():
			sheet_rows=[]
			ws=wb[sheet_name]
			if len(ws.rows)<2:
				continue
			for i in range(1,len(ws.rows)+1):
				sheet_row=[]
				for j in range(1,len(ws.columns)+1):
					cellvalue=ws.cell(row=i,column=j).value
					if cellvalue:
						cellvalue=str(cellvalue).strip()
					if i==1:
						for key,value in self.casedic.iteritems():
							if cellvalue==value:
								self.casedic[key]=j
					elif i>1 and ws.cell(row=i,column=self.casedic['runflag']).value!=0:
						sheet_row.append(cellvalue)
					else:
						break						
				if sheet_row and sheet_row!=[None]*len(sheet_row):
					sheet_rows.append(sheet_row)
			excel_rows.append([sheet_name,sheet_rows])
		return excel_rows


if __name__=='__main__':
    print 'No main file'
